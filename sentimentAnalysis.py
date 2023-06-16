# ! /usr/bin/python
# -*- coding: utf-8 -*-

# Improved import of modules and utilities
import sys
from os import path
from openpyxl import load_workbook
from paralleldots import set_api_key, sentiment
from time import sleep
from decouple import config
from constants import SENTIMENTS, DOC_HEADERS


class Analytics:

    # Use of environment variables, as good development practices
    def __init__(self, path, key=config('API_KEY')):
        self.path = path
        self.key = key

    # This method uses text analysis and allows the visualization in the xlsx file
    def process_file(self):
        if self.exist_file():
            set_api_key(self.key)
            workbook = load_workbook(self.path)
            sheet = workbook.active
            max_row = sheet.max_row

            # The defined constants are iterated to assign the headers to the values obtained in the text analysis
            for index, header in enumerate(DOC_HEADERS, start=4):
                sheet.cell(1, index).value = header

            # The prompts or texts are analyzed and the values are assigned in the cells for their display
            print('Processing...')
            for row in range(2, max_row + 1):
                prompt = str(sheet.cell(row, 3).value)

                output_sentiment = sentiment(prompt)

                """"
                    The enumerate function eliminates the redundancies in the code present in the
                    and a better efficiency in the assignment of cell values is achieved.
                """

                for index, item in enumerate(SENTIMENTS, start=4):
                    prompt_value_rounded = round(output_sentiment['sentiment'].get(
                        item) * 100, 3) if 'sentiment' in output_sentiment else 0
                        
                    sheet.cell(row, index).value = prompt_value_rounded

                sleep(10)
            workbook.save('sentimentAnalysis.xlsx')
            print('Successfully')
        else:
            print('Not exist file.')

    # This method verifies the existence of the file entered in the script arguments
    def exist_file(self):
        return path.isfile(self.path)


if __name__ == '__main__':
    if len(sys.argv) > 1:
        analitycs = Analytics(sys.argv[1]) if len(
            sys.argv) <= 2 else Analytics(sys.argv[2])
        analitycs.process_file()
    else:
        print("Select a file.")
